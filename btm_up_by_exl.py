import openpyxl
import pyodbc

col_start = 107
row_start = 14
base_rate = 0.65
labor_rate = 0.929
volume_db_name = "Volume_new"

car_List = ["VE0012", "VE0014", "VE0015", "VE0017", "VE0013", "VE0011", "VE0006", "VE0002",
            "VE0003", "VE0007", "VE0016", "VE0010", "VE0018", "VE0021", "VE0008", "VE0004",
            "VE0005", "VE0009", "VE0001", "VE0019", "VE0020"]

china_car_List = ["VE0016", "VE0010", "VE0018", "VE0021", "VE0019", "VE0020"]


def pre_proc(cursor, work_sht):

    if work_sht.cell(row=row_start-3, column=col_start).value == "ev12":

        for i in range(0,42):
            work_sht.cell(row=row_start-3, column=col_start+i).value = car_no_to_car_name(work_sht.cell(row=row_start-3, column=col_start+i).value)

    if work_sht.cell(row=row_start, column=2).value is None:

        for i in range(row_start, row_start + 391):
            module_name = work_sht.cell(row=i, column=7).value
            module_revision = work_sht.cell(row=i, column=6).value
            if module_revision is None: module_revision = ""
            if module_name is not None and module_revision.find("삭제") == -1 and module_revision.find("담당팀변경") == -1:
                module_name = str.replace(module_name, "'", "''")
                sQuery = "select ID from ModuleList where Item_name ='" + module_name + "'"
                cursor.execute(sQuery)
                q_list = list(cursor.fetchall())
                if len(q_list) > 0:
                    work_sht.cell(row=i, column=2).value = q_list[0][0]


def cell_input(shtname, seq, car_code, module_code, pai_code, value):

    shtname.cell(row=seq, column=1).value = car_code
    shtname.cell(row=seq, column=2).value = module_code
    shtname.cell(row=seq, column=3).value = pai_code
    shtname.cell(row=seq, column=4).value = value
    print(str(module_code) + "_" + str(pai_code) + "_" + car_code)
    shtname.cell(row=seq, column=5).value = module_code + "_" + pai_code + "_" + car_code

    return shtname


def is_non_local(is_strategic, is_universal):
    non_local = False

    if is_strategic == "●" or is_universal == "●":
        non_local = True

    return non_local


def list_chg(temp_list):
    output_list = []
    for entities in temp_list:
        output_list.append(entities[0])

    return output_list


def find_max(base_data_dict, car_list):
    max_value = 0
    for car_name in car_list:
        if base_data_dict[car_name] is None:
            base_data_dict[car_name] = 0
        if max_value < base_data_dict[car_name]:
            max_value = base_data_dict[car_name]

    # print("max value is " +str(max_value))
    return max_value


def cal_volume(volume_dict, car_list):
    volume = 0

    for car_name in car_list:
        volume = volume + volume_dict[car_name]

    # print("cal volume is "+str(volume))
    return volume


def cal_allocation(cursor, config_type, base_data_dict):

    selectQuery = "select sum(volume) from " + volume_db_name
    cursor.execute(selectQuery)
    q_list = list(cursor.fetchall())
    volume_total = q_list[0][0]
    car_volume_dict = {}

    # print("total volume = " + str(volume_total))

    for i in range(0, 21):
        selectQuery = "select sum(volume) from " + volume_db_name + " where ID='" + car_List[i] + "'"
        cursor.execute(selectQuery)
        q_list = list(cursor.fetchall())
        car_volume_dict[car_List[i]] = q_list[0][0]

    # print(car_volume_dict)

    selectQuery = "select ID from VehicleData where SEG='B' or SEG='C'"
    cursor.execute(selectQuery)
    c_seg_list = list_chg(cursor.fetchall())

    selectQuery = "select ID from VehicleData where SEG='D'"
    cursor.execute(selectQuery)
    d_seg_list = list_chg(cursor.fetchall())

    selectQuery = "select ID from VehicleData where SEG='E' or SEG='E+' "
    cursor.execute(selectQuery)
    e_seg_list = list_chg(cursor.fetchall())

    c_seg_volume = cal_volume(car_volume_dict, c_seg_list)
    d_seg_volume = cal_volume(car_volume_dict, d_seg_list)
    e_seg_volume = cal_volume(car_volume_dict, e_seg_list)

    selectQuery = "select ID from VehicleData where Brand='Hyundai' or Brand='Kia' "
    cursor.execute(selectQuery)
    hk_brand_list = list_chg(cursor.fetchall())

    selectQuery = "select ID from VehicleData where Brand='Genesis'"
    cursor.execute(selectQuery)
    g_brand_list = list_chg(cursor.fetchall())

    hk_brand_volume = cal_volume(car_volume_dict, hk_brand_list)
    g_brand_volume = cal_volume(car_volume_dict, g_brand_list)

    selectQuery = "select ID from VehicleData where BT='Hatchback' or BT='Sedan'"
    cursor.execute(selectQuery)
    sedan_list = list_chg(cursor.fetchall())

    selectQuery = "select ID from VehicleData where BT='CUV' or BT='SUV'"
    cursor.execute(selectQuery)
    suv_list = list_chg(cursor.fetchall())

    sedan_volume = cal_volume(car_volume_dict, sedan_list)
    suv_volume = cal_volume(car_volume_dict, suv_list)

    print("c_seg_volume : " + str(c_seg_volume) + "/d_seg_volume : " + str(d_seg_volume) + "/e_seg_volume : " + str(
        e_seg_volume))

    print("sedan_volume : " + str(sedan_volume) + "/suv_volume : " + str(suv_volume))

    if config_type == "통합":
        lead_car_value = find_max(base_data_dict, car_List)
        for car_name in base_data_dict.keys():
            base_data_dict[car_name] = lead_car_value * car_volume_dict[car_name] / volume_total

    elif config_type == "차급":

        c_lead_car_value = find_max(base_data_dict, c_seg_list)
        d_lead_car_value = find_max(base_data_dict, d_seg_list)
        e_lead_car_value = find_max(base_data_dict, e_seg_list)

        for c_seg_car_name in c_seg_list:
            base_data_dict[c_seg_car_name] = c_lead_car_value * car_volume_dict[c_seg_car_name] / c_seg_volume

        for d_seg_car_name in d_seg_list:
            base_data_dict[d_seg_car_name] = d_lead_car_value * car_volume_dict[d_seg_car_name] / d_seg_volume

        for e_seg_car_name in e_seg_list:
            base_data_dict[e_seg_car_name] = e_lead_car_value * car_volume_dict[e_seg_car_name] / e_seg_volume

    elif config_type == "브랜드":

        hk_lead_car_value = find_max(base_data_dict, hk_brand_list)
        g_lead_car_value = find_max(base_data_dict, g_brand_list)

        for h_brand_car_name in hk_brand_list:
            base_data_dict[h_brand_car_name] = hk_lead_car_value * car_volume_dict[h_brand_car_name] / hk_brand_volume

        for g_brand_car_name in g_brand_list:
            base_data_dict[g_brand_car_name] = g_lead_car_value * car_volume_dict[g_brand_car_name] / g_brand_volume

    elif config_type == "바디":

        sedan_lead_car_value = find_max(base_data_dict, sedan_list)
        suv_lead_car_value = find_max(base_data_dict, suv_list)

        for sedan_car_name in sedan_list:
            base_data_dict[sedan_car_name] = sedan_lead_car_value * car_volume_dict[sedan_car_name] / sedan_volume

        for suv_car_name in suv_list:
            base_data_dict[suv_car_name] = suv_lead_car_value * car_volume_dict[suv_car_name] / suv_volume

    return base_data_dict


def car_no_to_car_name(car_no):
    car_no = int(str.replace(car_no, "ev", ""))
    car_name = "VE00"
    if car_no < 10:
        car_name = car_name + "0" + str(car_no)
    if 10 <= car_no < 100:
        car_name = car_name + str(car_no)

    return car_name


def btm_up_chg(file_name,conn_String, sheet_name, UI_set):

    print("Start of allocating")

    targetBk = openpyxl.load_workbook(file_name, data_only=True)
    targetSht = targetBk[sheet_name]
    col_location = col_start

    while col_location < col_start + 41:

        UI_set.proc_bar.setValue(col_location)
        car_name = targetSht.cell(row=4, column=col_location).value

        for i in range(row_start, row_start + 391):

            non_local = is_non_local(targetSht.cell(row=i, column=9).value, targetSht.cell(row=i, column=10).value)

            # print(str(is_strategic) +" and "+str(is_universal))

            base_value = targetSht.cell(row=i, column=col_location).value
            app_value = targetSht.cell(row=i, column=col_location + 1).value

            # print(str(i) + " : " + str(is_allocated)+"base_value = "+str(base_value)+",app_value = "+str(app_value))

            if base_value is None:
                None

            elif base_value is not None and app_value is not None:
                None

            elif base_value is not None and app_value is None and non_local is True:
                targetSht.cell(row=i, column=col_location).value = base_value * base_rate
                targetSht.cell(row=i, column=col_location + 1).value = base_value * (1 - base_rate)

        col_location = col_location + 2

    for i in range(row_start, row_start + 391):

        non_local = is_non_local(targetSht.cell(row=i, column=9).value, targetSht.cell(row=i, column=10).value)

        col_location = col_start
        config_type = targetSht.cell(row=i, column=104).value
        base_data_dict = {}

        if non_local is True:

            while col_location < col_start + 41 :
                base_data_dict[targetSht.cell(row=4, column=col_location).value] = targetSht.cell(row=i,
                                                                                                  column=col_location).value
                col_location = col_location + 2

            # print(config_type +" : "+str(base_data_dict))
            base_data_dict = cal_allocation(conn_String, config_type, base_data_dict)
            # print(base_data_dict)
            col_location = col_start

            while col_location < col_start + 41:

                targetSht.cell(row=i, column=col_location).value = base_data_dict[
                    targetSht.cell(row=4, column=col_location).value]
                if config_type == "X":
                    targetSht.cell(row=i, column=col_location).value = 0
                    targetSht.cell(row=i, column=col_location + 1).value = 0
                col_location = col_location + 2

    targetBk.save(file_name)
    targetBk.close()

    print("End of allocating")


def check_china_car(car_name, china_car_list):
    is_china = False

    for china_car_name in china_car_list:
        if car_name == china_car_name:
            is_china = True

    return is_china


def data_sht_init(data_file_name):

    dataBk = openpyxl.load_workbook(data_file_name, read_only=False, keep_vba=True)
    hasDataSht = False

    for temp_sht_name in dataBk.sheetnames:
        temp_sht_name = "Data"
        hasDataSht = True

    if hasDataSht is not True:
        dataBk.create_sheet("Data")
        dataBk.cells(row=1, column=1).value = "ID"
        dataBk.cells(row=1, column=2).value = "Module"
        dataBk.cells(row=1, column=3).value = "PAI"
        dataBk.cells(row=1, column=4).value = "Value"

    dataSht = dataBk["Data"]
    dataSht.delete_rows(2, dataSht.max_row + 1)

    return dataBk


def btm_up_datacube(file_name, data_file_name, sheet_name, UI_file, cursor):

    print("Start of btm data cube making")
    #dataBk = data_sht_init(data_file_name)
    dataBk = openpyxl.load_workbook(data_file_name, read_only=False, keep_vba=True)
    dataSht= dataBk["Data"]
    dataSht.delete_rows(2, dataSht.max_row + 1)

    data_seq = 2

    targetBk = openpyxl.load_workbook(file_name, data_only=True)
    targetSht = targetBk[sheet_name]

    pre_proc(cursor, targetSht)

    col_location = col_start

    while col_location < col_start + 41:

        UI_file.proc_bar.setValue(col_location+1)
        car_name = targetSht.cell(row=row_start-3, column=col_location).value
        is_china = check_china_car(car_name, china_car_List)
        module_name = ""
        i = row_start

        while module_name != "MO0335":

            non_local = is_non_local(targetSht.cell(row=i, column=9).value, targetSht.cell(row=i, column=10).value)

            base_value = targetSht.cell(row=i, column=col_location).value

            if base_value is None:
                base_value = 0

            app_value = targetSht.cell(row=i, column=col_location + 1).value

            if app_value is None:
                app_value = 0

            module_name = targetSht.cell(row=i, column=2).value

            if module_name is not None and is_china is False:

                if non_local is True:
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI042", base_value)
                    dataSht = cell_input(dataSht, data_seq + 1, car_name, module_name, "NewPAI041",
                                         base_value * labor_rate)
                    dataSht = cell_input(dataSht, data_seq + 2, car_name, module_name, "NewPAI030", app_value)
                    dataSht = cell_input(dataSht, data_seq + 3, car_name, module_name, "NewPAI029",
                                         app_value * labor_rate)
                    data_seq = data_seq + 4
                else:
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI028", base_value)
                    dataSht = cell_input(dataSht, data_seq + 1, car_name, module_name, "NewPAI026",
                                         base_value * labor_rate)
                    data_seq = data_seq + 2

            elif module_name is not None and is_china is True:

                if non_local is True:
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI042", base_value)
                    dataSht = cell_input(dataSht, data_seq + 1, car_name, module_name, "NewPAI041",
                                         base_value * labor_rate)
                    dataSht = cell_input(dataSht, data_seq + 2, car_name, module_name, "NewPAI030", app_value)
                    dataSht = cell_input(dataSht, data_seq + 3, car_name, module_name, "NewPAI029",
                                         app_value * labor_rate)
                    data_seq = data_seq + 4
                else:
                    dataSht = cell_input(dataSht, data_seq, car_name, module_name, "NewPAI028", app_value)
                    dataSht = cell_input(dataSht, data_seq + 1, car_name, module_name, "NewPAI026",
                                         app_value * labor_rate)
                    data_seq = data_seq + 2

            i = i + 1

        col_location = col_location + 2

    targetBk.save(file_name)
    targetBk.close()
    dataBk.save(data_file_name)
    dataBk.close()
    print("End of btm data cube making")